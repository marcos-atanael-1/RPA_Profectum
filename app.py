from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import json
import subprocess
import threading
import os
from pathlib import Path
import sqlite3
import secrets
import string
import config

app = Flask(__name__)
app.config['SECRET_KEY'] = config.SECRET_KEY
app.config['SQLALCHEMY_DATABASE_URI'] = config.SQLALCHEMY_DATABASE_URI
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = config.SQLALCHEMY_TRACK_MODIFICATIONS
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
    'connect_args': {
        'timeout': 30,
        'check_same_thread': False
    }
}

db = SQLAlchemy(app)

# Configurar SQLite para melhor concorrência (WAL mode)
from sqlalchemy import text

@app.before_request
def setup_database():
    if not hasattr(app, 'db_configured'):
        try:
            with db.engine.connect() as conn:
                conn.execute(text('PRAGMA journal_mode=WAL'))
                conn.execute(text('PRAGMA busy_timeout=30000'))
                conn.commit()
            app.db_configured = True
        except:
            pass  # Ignora se já estiver configurado
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor, faça login para acessar esta página.'
login_manager.login_message_category = 'info'

# Modelos do banco de dados
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    full_name = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='user')  # admin, user
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)
    reset_token = db.Column(db.String(100))
    reset_token_expires = db.Column(db.DateTime)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    def generate_reset_token(self):
        self.reset_token = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(32))
        self.reset_token_expires = datetime.utcnow() + timedelta(hours=1)
        return self.reset_token
    
    def is_admin(self):
        return self.role == 'admin'
    
    def to_dict(self):
        return {
            'id': self.id,
            'username': self.username,
            'email': self.email,
            'full_name': self.full_name,
            'role': self.role,
            'is_active': self.is_active,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'last_login': self.last_login.isoformat() if self.last_login else None
        }

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

class SystemSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.Text)
    description = db.Column(db.String(500))
    updated_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    @staticmethod
    def get_setting(key, default_value=None):
        setting = SystemSettings.query.filter_by(key=key).first()
        return setting.value if setting else default_value
    
    @staticmethod
    def set_setting(key, value, description=None, user_id=None):
        setting = SystemSettings.query.filter_by(key=key).first()
        if setting:
            setting.value = value
            setting.updated_by = user_id
            setting.updated_at = datetime.utcnow()
        else:
            setting = SystemSettings(
                key=key,
                value=value,
                description=description,
                updated_by=user_id
            )
            db.session.add(setting)
        db.session.commit()
        return setting

class RecebimentoNF(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    pedido_compra = db.Column(db.String(100), nullable=False)
    nota_fiscal = db.Column(db.String(100), nullable=False)
    chave_acesso = db.Column(db.String(44), nullable=False)  # Chave NFe tem 44 caracteres
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    status = db.Column(db.String(20), nullable=False, default='pendente')  # pendente, processado, erro
    error_message = db.Column(db.Text)  # Mensagem de erro caso ocorra
    
    creator = db.relationship('User', backref=db.backref('recebimentos_nf', lazy=True))
    
    def to_dict(self):
        return {
            'id': self.id,
            'pedido_compra': self.pedido_compra,
            'nota_fiscal': self.nota_fiscal,
            'chave_acesso': self.chave_acesso,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'created_by': self.created_by,
            'creator_name': self.creator.full_name if self.creator else None,
            'status': self.status,
            'error_message': self.error_message
        }

class BotExecution(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bot_name = db.Column(db.String(100), nullable=False)
    status = db.Column(db.String(20), nullable=False)  # running, completed, failed, stopped
    start_time = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    end_time = db.Column(db.DateTime)
    duration = db.Column(db.Float)  # em segundos
    parameters = db.Column(db.Text)  # JSON string
    result = db.Column(db.Text)
    error_message = db.Column(db.Text)
    
    def to_dict(self):
        return {
            'id': self.id,
            'bot_name': self.bot_name,
            'status': self.status,
            'start_time': self.start_time.isoformat() if self.start_time else None,
            'end_time': self.end_time.isoformat() if self.end_time else None,
            'duration': self.duration,
            'parameters': self.parameters,
            'result': self.result,
            'error_message': self.error_message
        }

class BotLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    execution_id = db.Column(db.Integer, db.ForeignKey('bot_execution.id'), nullable=False)
    timestamp = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    level = db.Column(db.String(20), nullable=False)  # INFO, WARNING, ERROR, DEBUG
    message = db.Column(db.Text, nullable=False)
    module = db.Column(db.String(100))  # qual módulo gerou o log
    
    execution = db.relationship('BotExecution', backref=db.backref('logs', lazy=True))
    
    def to_dict(self):
        return {
            'id': self.id,
            'execution_id': self.execution_id,
            'timestamp': self.timestamp.isoformat(),
            'level': self.level,
            'message': self.message,
            'module': self.module
        }

# ============================================================================
# MODELOS DE ROMANEIOS
# ============================================================================

class Romaneio(db.Model):
    """Modelo para gerenciar romaneios/pedidos de compra"""
    __tablename__ = 'romaneio'
    
    id = db.Column(db.Integer, primary_key=True)
    pedido_compra = db.Column(db.String(100), nullable=False, unique=True, index=True)
    nota_fiscal = db.Column(db.String(100), nullable=False)
    chave_acesso = db.Column(db.String(44), nullable=False)
    idro = db.Column(db.Integer, nullable=True)
    status = db.Column(db.String(1), nullable=False, default='P', index=True)
    tentativas_contagem = db.Column(db.Integer, nullable=False, default=0)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, onupdate=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    creator = db.relationship('User', backref=db.backref('romaneios', lazy=True))
    observacoes = db.Column(db.Text, nullable=True)
    apos_recebimento = db.Column(db.Boolean, default=False)
    programado = db.Column(db.Boolean, default=True)
    inserir_como_parcial = db.Column(db.Boolean, default=False)
    
    itens = db.relationship('RomaneioItem', backref='romaneio', lazy=True, cascade='all, delete-orphan')
    logs = db.relationship('RomaneioLog', backref='romaneio', lazy=True, cascade='all, delete-orphan')
    
    def to_dict(self):
        return {
            'id': self.id,
            'pedido_compra': self.pedido_compra,
            'nota_fiscal': self.nota_fiscal,
            'chave_acesso': self.chave_acesso,
            'idro': self.idro,
            'status': self.status,
            'status_label': config.STATUS_CHOICES.get(self.status, 'Desconhecido'),
            'status_color': config.STATUS_COLORS.get(self.status, 'secondary'),
            'tentativas_contagem': self.tentativas_contagem,
            'max_tentativas': config.MAX_TENTATIVAS_CONTAGEM,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None,
            'created_by': self.created_by,
            'creator_name': self.creator.full_name if self.creator else None,
            'observacoes': self.observacoes,
            'apos_recebimento': self.apos_recebimento,
            'programado': self.programado,
            'inserir_como_parcial': self.inserir_como_parcial,
            'total_itens': len(self.itens),
            'itens_divergentes': sum(1 for item in self.itens if item.tem_divergencia())
        }
    
    def pode_excluir(self):
        return self.status == 'P' and self.tentativas_contagem == 0
    
    def pode_verificar(self):
        return self.status not in ['F'] and self.tentativas_contagem < config.MAX_TENTATIVAS_CONTAGEM
    
    def incrementar_tentativa(self):
        self.tentativas_contagem += 1
        self.updated_at = datetime.utcnow()

class RomaneioItem(db.Model):
    """Itens de um romaneio"""
    __tablename__ = 'romaneio_item'
    
    id = db.Column(db.Integer, primary_key=True)
    romaneio_id = db.Column(db.Integer, db.ForeignKey('romaneio.id'), nullable=False, index=True)
    idro = db.Column(db.Integer, nullable=True)
    codigo = db.Column(db.String(50), nullable=False)
    descricao = db.Column(db.String(500), nullable=False)
    quantidade_nf = db.Column(db.Integer, nullable=False)
    quantidade_contada = db.Column(db.Integer, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def tem_divergencia(self):
        if self.quantidade_contada is None:
            return True
        return self.quantidade_nf != self.quantidade_contada
    
    def to_dict(self):
        return {
            'id': self.id,
            'romaneio_id': self.romaneio_id,
            'idro': self.idro,
            'codigo': self.codigo,
            'descricao': self.descricao,
            'quantidade_nf': self.quantidade_nf,
            'quantidade_contada': self.quantidade_contada,
            'tem_divergencia': self.tem_divergencia(),
            'divergencia_valor': (self.quantidade_contada - self.quantidade_nf) if self.quantidade_contada else None,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }

class RomaneioLog(db.Model):
    """Log de mudanças e ações em romaneios"""
    __tablename__ = 'romaneio_log'
    
    id = db.Column(db.Integer, primary_key=True)
    romaneio_id = db.Column(db.Integer, db.ForeignKey('romaneio.id'), nullable=False, index=True)
    timestamp = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, index=True)
    acao = db.Column(db.String(50), nullable=False)
    status_anterior = db.Column(db.String(1), nullable=True)
    status_novo = db.Column(db.String(1), nullable=True)
    tentativa = db.Column(db.Integer, nullable=True)
    detalhes = db.Column(db.Text, nullable=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    user = db.relationship('User', backref=db.backref('romaneio_logs', lazy=True))
    
    def to_dict(self):
        return {
            'id': self.id,
            'romaneio_id': self.romaneio_id,
            'timestamp': self.timestamp.isoformat() if self.timestamp else None,
            'acao': self.acao,
            'status_anterior': self.status_anterior,
            'status_anterior_label': config.STATUS_CHOICES.get(self.status_anterior, '') if self.status_anterior else None,
            'status_novo': self.status_novo,
            'status_novo_label': config.STATUS_CHOICES.get(self.status_novo, '') if self.status_novo else None,
            'tentativa': self.tentativa,
            'detalhes': self.detalhes,
            'user_id': self.user_id,
            'user_name': self.user.full_name if self.user else 'Sistema Automático'
        }

# Configuração dos bots disponíveis
AVAILABLE_BOTS = {
    'sic_full': {
        'name': 'SIC - Processo Completo',
        'description': 'Login no SIC + Acesso ao módulo contábil/fiscal',
        'script': 'entrada-nf/bot.py',
        'estimated_duration': 300  # segundos
    },
    'sic_login': {
        'name': 'SIC - Apenas Login',
        'description': 'Realiza apenas o login no sistema SIC',
        'script': 'entrada-nf/Sic_Login.py',
        'estimated_duration': 60
    },
    'sic_inserir_nfs': {
        'name': 'SIC - Inserir NFs Pendentes',
        'description': 'Busca todas as NFs com status pendente no banco e insere no sistema SIC',
        'script': 'entrada-nf/Sic_Inserir_NF.py',
        'estimated_duration': 180
    },
    'rm_login': {
        'name': 'RM - Login',
        'description': 'Realiza login no sistema TOTVS RM',
        'script': 'entrada-nf/RM_Login.py',
        'estimated_duration': 60
    },
    'consulta_nfe': {
        'name': 'Consulta NFe',
        'description': 'Consulta nota fiscal eletrônica via API',
        'script': 'entrada-nf/Consulta_nfe.py',
        'estimated_duration': 30
    }
}

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login"""
    if current_user.is_authenticated:
        return redirect(url_for('romaneios'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = bool(request.form.get('remember'))
        
        if not username or not password:
            flash('Por favor, preencha todos os campos.', 'error')
            return render_template('auth/login.html')
        
        user = User.query.filter((User.username == username) | (User.email == username)).first()
        
        if user and user.check_password(password) and user.is_active:
            user.last_login = datetime.utcnow()
            db.session.commit()
            login_user(user, remember=remember)
            
            # Redirect to intended page or dashboard
            next_page = request.args.get('next')
            if next_page:
                return redirect(next_page)
            return redirect(url_for('romaneios'))
        else:
            flash('Usuário ou senha inválidos.', 'error')
    
    return render_template('auth/login.html')

@app.route('/logout')
@login_required
def logout():
    """Logout do usuário"""
    logout_user()
    flash('Logout realizado com sucesso.', 'success')
    return redirect(url_for('login'))

@app.route('/register', methods=['POST'])
@login_required
def register():
    """Cadastro de usuários via modal (apenas admins)"""
    if not current_user.is_admin():
        flash('Acesso negado. Apenas administradores podem cadastrar usuários.', 'error')
        return redirect(url_for('romaneios'))
    
    username = request.form.get('username', '').strip()
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '').strip()
    full_name = request.form.get('full_name', '').strip()
    role = request.form.get('role', 'user')
    
    # Validações
    if not all([username, email, password, full_name]):
        flash('Por favor, preencha todos os campos.', 'error')
        return redirect(url_for('users_management'))
    
    if len(password) < 6:
        flash('A senha deve ter pelo menos 6 caracteres.', 'error')
        return redirect(url_for('users_management'))
    
    # Verificar se usuário já existe
    if User.query.filter_by(username=username).first():
        flash('Nome de usuário já existe.', 'error')
        return redirect(url_for('users_management'))
    
    if User.query.filter_by(email=email).first():
        flash('E-mail já está em uso.', 'error')
        return redirect(url_for('users_management'))
    
    # Criar novo usuário
    try:
        user = User(
            username=username,
            email=email,
            full_name=full_name,
            role=role
        )
        user.set_password(password)
        
        db.session.add(user)
        db.session.flush()  # Flush antes do commit
        db.session.commit()
        
        flash(f'✅ Usuário {username} cadastrado com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        import traceback
        error_msg = str(e)
        # Se for erro de database locked, dar uma dica
        if 'database is locked' in error_msg:
            flash('⚠️ Banco de dados ocupado. Por favor, tente novamente em alguns segundos.', 'warning')
        else:
            flash(f'Erro ao criar usuário: {error_msg}', 'error')
        print(f"Erro detalhado ao criar usuário: {traceback.format_exc()}")
    finally:
        db.session.close()
    
    return redirect(url_for('users_management'))

@app.route('/users')
@login_required
def users_management():
    """Página de gerenciamento de usuários (apenas admins)"""
    if not current_user.is_admin():
        flash('Acesso negado. Apenas administradores podem gerenciar usuários.', 'error')
        return redirect(url_for('romaneios'))
    
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('auth/users.html', users=users)

@app.route('/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
def admin_reset_user_password(user_id):
    """Admin reseta senha de um usuário"""
    if not current_user.is_admin():
        return jsonify({'error': 'Acesso negado'}), 403
    
    user = User.query.get_or_404(user_id)
    
    if user.id == current_user.id:
        return jsonify({'error': 'Você não pode resetar sua própria senha por aqui'}), 400
    
    data = request.get_json()
    new_password = data.get('new_password', '').strip()
    
    if not new_password or len(new_password) < 6:
        return jsonify({'error': 'Senha deve ter no mínimo 6 caracteres'}), 400
    
    user.set_password(new_password)
    db.session.commit()
    
    return jsonify({'message': f'Senha do usuário {user.username} resetada com sucesso!'})

@app.route('/users/<int:user_id>/toggle-status', methods=['POST'])
@login_required
def toggle_user_status(user_id):
    """Ativar/desativar um usuário"""
    if not current_user.is_admin():
        return jsonify({'error': 'Acesso negado'}), 403
    
    user = User.query.get_or_404(user_id)
    
    if user.id == current_user.id:
        return jsonify({'error': 'Você não pode desativar sua própria conta'}), 400
    
    user.is_active = not user.is_active
    action = 'ativado' if user.is_active else 'desativado'
    db.session.commit()
    
    return jsonify({'message': f'Usuário {user.username} {action} com sucesso!'})

@app.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
def delete_user(user_id):
    """Excluir um usuário"""
    if not current_user.is_admin():
        return jsonify({'error': 'Acesso negado'}), 403
    
    user = User.query.get_or_404(user_id)
    
    if user.id == current_user.id:
        return jsonify({'error': 'Você não pode excluir sua própria conta'}), 400
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    
    return jsonify({'message': f'Usuário {username} excluído com sucesso!'})

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Usuário troca sua própria senha"""
    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Validações
        if not current_password:
            flash('Senha atual é obrigatória', 'error')
            return render_template('auth/change_password.html')
        
        if not current_user.check_password(current_password):
            flash('Senha atual incorreta', 'error')
            return render_template('auth/change_password.html')
        
        if not new_password or len(new_password) < 6:
            flash('A nova senha deve ter no mínimo 6 caracteres', 'error')
            return render_template('auth/change_password.html')
        
        if new_password != confirm_password:
            flash('As senhas não coincidem', 'error')
            return render_template('auth/change_password.html')
        
        if current_password == new_password:
            flash('A nova senha deve ser diferente da senha atual', 'error')
            return render_template('auth/change_password.html')
        
        # Alterar senha
        current_user.set_password(new_password)
        db.session.commit()
        
        flash('Senha alterada com sucesso!', 'success')
        return redirect(url_for('romaneios'))
    
    return render_template('auth/change_password.html')

@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password_request():
    """Solicitar reset de senha"""
    if current_user.is_authenticated:
        return redirect(url_for('romaneios'))
    
    if request.method == 'POST':
        email = request.form.get('email')
        
        if not email:
            flash('Por favor, digite seu e-mail.', 'error')
            return render_template('auth/reset_request.html')
        
        user = User.query.filter_by(email=email).first()
        
        if user and user.is_active:
            token = user.generate_reset_token()
            db.session.commit()
            
            # Aqui você implementaria o envio de e-mail
            # Por enquanto, vamos mostrar o token na tela (apenas para desenvolvimento)
            flash(f'Token de reset gerado: {token} (válido por 1 hora)', 'info')
            return redirect(url_for('reset_password_with_token', token=token))
        else:
            flash('E-mail não encontrado.', 'error')
    
    return render_template('auth/reset_request.html')

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password_with_token(token):
    """Reset de senha com token"""
    if current_user.is_authenticated:
        return redirect(url_for('romaneios'))
    
    user = User.query.filter_by(reset_token=token).first()
    
    if not user or not user.reset_token_expires or user.reset_token_expires < datetime.utcnow():
        flash('Token inválido ou expirado.', 'error')
        return redirect(url_for('reset_password_request'))
    
    if request.method == 'POST':
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if not password or not confirm_password:
            flash('Por favor, preencha todos os campos.', 'error')
            return render_template('auth/reset_password.html', token=token)
        
        if password != confirm_password:
            flash('As senhas não coincidem.', 'error')
            return render_template('auth/reset_password.html', token=token)
        
        if len(password) < 6:
            flash('A senha deve ter pelo menos 6 caracteres.', 'error')
            return render_template('auth/reset_password.html', token=token)
        
        user.set_password(password)
        user.reset_token = None
        user.reset_token_expires = None
        db.session.commit()
        
        flash('Senha alterada com sucesso! Faça login com sua nova senha.', 'success')
        return redirect(url_for('login'))
    
    return render_template('auth/reset_password.html', token=token)

@app.route('/settings')
@login_required
def settings():
    if not current_user.is_admin():
        flash('Acesso negado. Apenas administradores podem acessar as configurações.', 'error')
        return redirect(url_for('romaneios'))
    
    # Buscar configurações atuais
    current_settings = {
        'primary_color': SystemSettings.get_setting('primary_color', '#2563eb'),
        'logo_url': SystemSettings.get_setting('logo_url', ''),
        'logo_type': SystemSettings.get_setting('logo_type', 'icon'),  # icon, image, url
        'system_name': SystemSettings.get_setting('system_name', 'RPA Profectum'),
        # Configurações de login
        'login_bg_type': SystemSettings.get_setting('login_bg_type', 'color'),
        'login_bg_color': SystemSettings.get_setting('login_bg_color', '#f8fafc'),
        'login_gradient_color1': SystemSettings.get_setting('login_gradient_color1', '#2563eb'),
        'login_gradient_color2': SystemSettings.get_setting('login_gradient_color2', '#1d4ed8'),
        'login_bg_image_url': SystemSettings.get_setting('login_bg_image_url', ''),
        'login_height': SystemSettings.get_setting('login_height', 'normal'),
        'login_title': SystemSettings.get_setting('login_title', 'Bem-vindo de volta!'),
    }
    
    return render_template('settings.html', settings=current_settings)

@app.route('/settings', methods=['POST'])
@login_required
def update_settings():
    if not current_user.is_admin():
        flash('Acesso negado. Apenas administradores podem modificar as configurações.', 'error')
        return redirect(url_for('romaneios'))
    
    try:
        settings_type = request.form.get('settings_type', 'system')
        
        if settings_type == 'login':
            # Configurações de login
            SystemSettings.set_setting('login_bg_type', request.form.get('login_bg_type', 'color'), 'Tipo de fundo da página de login', current_user.id)
            SystemSettings.set_setting('login_bg_color', request.form.get('login_bg_color', '#f8fafc'), 'Cor de fundo da página de login', current_user.id)
            SystemSettings.set_setting('login_gradient_color1', request.form.get('login_gradient_color1', '#2563eb'), 'Primeira cor do gradiente de login', current_user.id)
            SystemSettings.set_setting('login_gradient_color2', request.form.get('login_gradient_color2', '#1d4ed8'), 'Segunda cor do gradiente de login', current_user.id)
            SystemSettings.set_setting('login_bg_image_url', request.form.get('login_bg_image_url', ''), 'URL da imagem de fundo do login', current_user.id)
            SystemSettings.set_setting('login_height', request.form.get('login_height', 'normal'), 'Altura da página de login', current_user.id)
            SystemSettings.set_setting('login_title', request.form.get('login_title', 'Bem-vindo de volta!'), 'Título da página de login', current_user.id)
            flash('Configurações de login atualizadas com sucesso!', 'success')
        else:
            # Configurações do sistema
            primary_color = request.form.get('primary_color', '#2563eb')
            SystemSettings.set_setting('primary_color', primary_color, 'Cor primária do sistema', current_user.id)
            
            system_name = request.form.get('system_name', 'RPA Profectum')
            SystemSettings.set_setting('system_name', system_name, 'Nome do sistema', current_user.id)
            
            logo_type = request.form.get('logo_type', 'icon')
            SystemSettings.set_setting('logo_type', logo_type, 'Tipo de logo (icon/image/url)', current_user.id)
            
            if logo_type in ['image', 'url']:
                logo_url = request.form.get('logo_url', '')
                SystemSettings.set_setting('logo_url', logo_url, 'URL da logo personalizada', current_user.id)
            
            flash('Configurações do sistema atualizadas com sucesso!', 'success')
        
    except Exception as e:
        flash(f'Erro ao atualizar configurações: {str(e)}', 'error')
    
    return redirect(url_for('settings'))

@app.route('/')
@login_required
def index():
    """Redireciona para a página principal de romaneios"""
    return redirect(url_for('romaneios'))



@app.route('/logs')
@login_required
def logs_view():
    """Página de visualização de logs de romaneios"""
    page = request.args.get('page', 1, type=int)
    
    # Filtros para logs de romaneio
    romaneio_id = request.args.get('romaneio_id', type=int)
    acao = request.args.get('acao', '')
    
    # Query de Logs de Romaneio
    romaneio_query = RomaneioLog.query
    if romaneio_id:
        romaneio_query = romaneio_query.filter_by(romaneio_id=romaneio_id)
    if acao:
        romaneio_query = romaneio_query.filter_by(acao=acao)
    
    romaneio_logs = romaneio_query.order_by(RomaneioLog.timestamp.desc()).paginate(
        page=page, per_page=50, error_out=False
    )
    
    romaneios = Romaneio.query.order_by(Romaneio.created_at.desc()).limit(50).all()
    
    return render_template('logs.html', 
                         romaneio_logs=romaneio_logs, 
                         romaneios=romaneios,
                         current_romaneio_id=romaneio_id,
                         current_acao=acao)

@app.route('/execute/<bot_id>', methods=['POST'])
@login_required
def execute_bot(bot_id):
    """Executa um bot específico"""
    if bot_id not in AVAILABLE_BOTS:
        return jsonify({'error': 'Bot não encontrado'}), 404
    
    bot_config = AVAILABLE_BOTS[bot_id]
    parameters = request.json.get('parameters', {})
    
    # Criar registro de execução
    execution = BotExecution(
        bot_name=bot_config['name'],
        status='running',
        parameters=json.dumps(parameters)
    )
    db.session.add(execution)
    db.session.commit()
    
    # Log inicial
    log_entry = BotLog(
        execution_id=execution.id,
        level='INFO',
        message=f'Iniciando execução do bot {bot_config["name"]}',
        module='orchestrator'
    )
    db.session.add(log_entry)
    db.session.commit()
    
    # Executar bot em thread separada
    def run_bot():
        try:
            # Executar o script do bot
            script_path = Path(bot_config['script'])
            if script_path.exists():
                # Configurar variáveis de ambiente para o bot
                env = os.environ.copy()
                env['RPA_EXECUTION_ID'] = str(execution.id)
                
                result = subprocess.run(
                    ['python', str(script_path)],
                    capture_output=True,
                    text=True,
                    cwd=Path.cwd(),
                    env=env
                )
                
                # Atualizar execução
                execution.status = 'completed' if result.returncode == 0 else 'failed'
                execution.end_time = datetime.utcnow()
                execution.duration = (execution.end_time - execution.start_time).total_seconds()
                execution.result = result.stdout
                if result.stderr:
                    execution.error_message = result.stderr
                
                # Log final
                log_level = 'INFO' if result.returncode == 0 else 'ERROR'
                log_message = 'Execução concluída com sucesso' if result.returncode == 0 else f'Execução falhou: {result.stderr}'
                
                final_log = BotLog(
                    execution_id=execution.id,
                    level=log_level,
                    message=log_message,
                    module='orchestrator'
                )
                db.session.add(final_log)
            else:
                execution.status = 'failed'
                execution.end_time = datetime.utcnow()
                execution.duration = (execution.end_time - execution.start_time).total_seconds()
                execution.error_message = f'Script não encontrado: {script_path}'
                
                error_log = BotLog(
                    execution_id=execution.id,
                    level='ERROR',
                    message=f'Script não encontrado: {script_path}',
                    module='orchestrator'
                )
                db.session.add(error_log)
            
            db.session.commit()
            
        except Exception as e:
            execution.status = 'failed'
            execution.end_time = datetime.utcnow()
            execution.duration = (execution.end_time - execution.start_time).total_seconds()
            execution.error_message = str(e)
            
            error_log = BotLog(
                execution_id=execution.id,
                level='ERROR',
                message=f'Erro durante execução: {str(e)}',
                module='orchestrator'
            )
            db.session.add(error_log)
            db.session.commit()
    
    # Iniciar thread
    thread = threading.Thread(target=run_bot)
    thread.daemon = True
    thread.start()
    
    return jsonify({
        'execution_id': execution.id,
        'message': f'Bot {bot_config["name"]} iniciado com sucesso',
        'estimated_duration': bot_config['estimated_duration']
    })

@app.route('/execution/<int:execution_id>')
@login_required
def execution_details(execution_id):
    """Detalhes de uma execução específica"""
    execution = BotExecution.query.get_or_404(execution_id)
    logs = BotLog.query.filter_by(execution_id=execution_id).order_by(BotLog.timestamp.asc()).all()
    
    return render_template('execution_details.html', execution=execution, logs=logs)

@app.route('/api/execution/<int:execution_id>/status')
def execution_status(execution_id):
    """API para verificar status de execução"""
    execution = BotExecution.query.get_or_404(execution_id)
    return jsonify(execution.to_dict())

@app.route('/api/logs/<int:execution_id>')
def execution_logs_api(execution_id):
    """API para obter logs de uma execução"""
    logs = BotLog.query.filter_by(execution_id=execution_id).order_by(BotLog.timestamp.asc()).all()
    return jsonify([log.to_dict() for log in logs])

@app.route('/stop/<int:execution_id>', methods=['POST'])
def stop_execution(execution_id):
    """Para uma execução em andamento"""
    execution = BotExecution.query.get_or_404(execution_id)
    
    if execution.status == 'running':
        execution.status = 'stopped'
        execution.end_time = datetime.utcnow()
        execution.duration = (execution.end_time - execution.start_time).total_seconds()
        
        stop_log = BotLog(
            execution_id=execution.id,
            level='WARNING',
            message='Execução interrompida pelo usuário',
            module='orchestrator'
        )
        db.session.add(stop_log)
        db.session.commit()
        
        flash('Execução interrompida com sucesso', 'warning')
    else:
        flash('Execução não está em andamento', 'error')
    
    return redirect(url_for('execution_details', execution_id=execution_id))

# ============================================================================
# ROTAS DE ROMANEIOS
# ============================================================================

@app.route('/romaneios')
@login_required
def romaneios():
    """Página de gerenciamento de romaneios"""
    page = request.args.get('page', 1, type=int)
    per_page = 15
    
    # Filtros
    status_filter = request.args.get('status', '')
    pedido_filter = request.args.get('pedido', '')
    nf_filter = request.args.get('nf', '')
    
    # Query base
    query = Romaneio.query
    
    # Aplicar filtros
    if status_filter:
        query = query.filter(Romaneio.status == status_filter)
    if pedido_filter:
        query = query.filter(Romaneio.pedido_compra.contains(pedido_filter))
    if nf_filter:
        query = query.filter(Romaneio.nota_fiscal.contains(nf_filter))
    
    # Ordenar por data de criação (mais recente primeiro)
    query = query.order_by(Romaneio.created_at.desc())
    
    # Paginação
    romaneios_paginados = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Estatísticas simples
    stats = {
        'total': Romaneio.query.count(),
        'pendentes': Romaneio.query.filter_by(status='P').count(),
        'abertos': Romaneio.query.filter_by(status='A').count(),
        'recebidos': Romaneio.query.filter_by(status='R').count(),
        'finalizados': Romaneio.query.filter_by(status='F').count()
    }
    
    return render_template('romaneios/lista.html',
                         romaneios=romaneios_paginados,
                         stats=stats,
                         status_filter=status_filter,
                         pedido_filter=pedido_filter,
                         nf_filter=nf_filter,
                         status_choices=config.STATUS_CHOICES,
                         modo_teste=config.MODO_TESTE)

@app.route('/romaneios/exportar-excel')
@login_required
def exportar_romaneios_excel():
    """Exportar romaneios para Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    # Filtros (mesmos da listagem)
    status_filter = request.args.get('status', '')
    pedido_filter = request.args.get('pedido', '')
    nf_filter = request.args.get('nf', '')
    
    # Query base
    query = Romaneio.query
    
    # Aplicar filtros
    if status_filter:
        query = query.filter(Romaneio.status == status_filter)
    if pedido_filter:
        query = query.filter(Romaneio.pedido_compra.contains(pedido_filter))
    if nf_filter:
        query = query.filter(Romaneio.nota_fiscal.contains(nf_filter))
    
    # Ordenar por data de criação (mais recente primeiro)
    romaneios = query.order_by(Romaneio.created_at.desc()).all()
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Romaneios"
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Cabeçalhos
    headers = [
        "Pedido de Compra", "Nota Fiscal", "Chave de Acesso", "Status",
        "Tentativas", "Total Itens", "Itens Divergentes", "IDRO",
        "Criado em", "Criado por", "Atualizado em"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Dados
    for row_num, romaneio in enumerate(romaneios, 2):
        dados = romaneio.to_dict()
        
        ws.cell(row=row_num, column=1).value = romaneio.pedido_compra
        ws.cell(row=row_num, column=2).value = romaneio.nota_fiscal
        ws.cell(row=row_num, column=3).value = romaneio.chave_acesso
        ws.cell(row=row_num, column=4).value = dados['status_label']
        ws.cell(row=row_num, column=5).value = f"{romaneio.tentativas_contagem}/{dados['max_tentativas']}"
        ws.cell(row=row_num, column=6).value = dados['total_itens']
        ws.cell(row=row_num, column=7).value = dados['itens_divergentes']
        ws.cell(row=row_num, column=8).value = romaneio.idro or '-'
        ws.cell(row=row_num, column=9).value = romaneio.created_at.strftime('%d/%m/%Y %H:%M') if romaneio.created_at else '-'
        ws.cell(row=row_num, column=10).value = romaneio.creator.full_name if romaneio.creator else '-'
        ws.cell(row=row_num, column=11).value = romaneio.updated_at.strftime('%d/%m/%Y %H:%M') if romaneio.updated_at else '-'
    
    # Ajustar largura das colunas
    column_widths = [18, 15, 48, 12, 12, 12, 18, 12, 18, 20, 18]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nome do arquivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'romaneios_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/romaneios/<int:romaneio_id>')
@login_required
def romaneio_detalhes(romaneio_id):
    """Página de detalhes de um romaneio específico"""
    romaneio = Romaneio.query.get_or_404(romaneio_id)
    
    # Buscar logs
    logs = RomaneioLog.query.filter_by(romaneio_id=romaneio_id)\
        .order_by(RomaneioLog.timestamp.desc()).all()
    
    return render_template('romaneios/detalhes.html',
                         romaneio=romaneio,
                         logs=logs,
                         status_choices=config.STATUS_CHOICES,
                         modo_teste=config.MODO_TESTE)

@app.route('/romaneios/add', methods=['POST'])
@login_required
def add_romaneio():
    """Adicionar novo romaneio"""
    from services.api_client import RomaneioAPIClient
    
    try:
        pedido_compra = request.form.get('pedido_compra', '').strip()
        nota_fiscal = request.form.get('nota_fiscal', '').strip()
        chave_acesso = request.form.get('chave_acesso', '').strip()
        
        # Validações
        if not pedido_compra:
            flash('Pedido de Compra é obrigatório', 'error')
            return redirect(url_for('romaneios'))
        if not nota_fiscal:
            flash('Nota Fiscal é obrigatória', 'error')
            return redirect(url_for('romaneios'))
        if not chave_acesso:
            flash('Chave de Acesso é obrigatória', 'error')
            return redirect(url_for('romaneios'))
        if len(chave_acesso) != 44:
            flash('Chave de Acesso deve ter 44 caracteres', 'error')
            return redirect(url_for('romaneios'))
        
        # Verificar se já existe
        existing = Romaneio.query.filter_by(pedido_compra=pedido_compra).first()
        if existing:
            flash('Já existe um romaneio para este pedido', 'error')
            return redirect(url_for('romaneios'))
        
        # Criar romaneio com valores padrão do .env
        romaneio = Romaneio(
            pedido_compra=pedido_compra,
            nota_fiscal=nota_fiscal,
            chave_acesso=chave_acesso,
            created_by=current_user.id,
            status='P',
            apos_recebimento=config.API_APOS_RECEBIMENTO,
            programado=config.API_PROGRAMADO,
            inserir_como_parcial=config.API_INSERIR_PARCIAL_SE_EXISTIR
        )
        
        # Chamar API se não for modo teste
        if not config.MODO_TESTE:
            api_client = RomaneioAPIClient()
            try:
                print("\n" + "="*80)
                print("🚀 CHAMANDO API DE INSERÇÃO DE ROMANEIO")
                print("="*80)
                print(f"📋 Pedido: {pedido_compra}")
                print(f"📄 Nota Fiscal: {nota_fiscal}")
                print(f"🔑 Chave: {chave_acesso}")
                print("-"*80)
                
                resultado_api = api_client.inserir_romaneio(
                    pedido_compra, nota_fiscal, chave_acesso
                )
                
                print("\n✅ RESPOSTA DA API:")
                print("-"*80)
                print(f"📊 Status: SUCCESS")
                print(f"📦 Response completo:")
                import json
                print(json.dumps(resultado_api, indent=2, ensure_ascii=False))
                print("="*80 + "\n")
                
                # Verificar se a API retornou mensagem de romaneio já existente
                if isinstance(resultado_api, dict):
                    mensagem = resultado_api.get('mensagem', '')
                    
                    # Verificar se é um erro de romaneio já existente
                    if 'já existente' in mensagem.lower() or 'ja existente' in mensagem.lower():
                        print(f"⚠️ Romaneio já existe na API externa!")
                        flash(f'❌ {mensagem}', 'error')
                        return redirect(url_for('romaneios'))
                    
                    # Se tiver IDRO, atualizar
                    if 'idro' in resultado_api:
                        romaneio.idro = resultado_api['idro']
                        print(f"✓ IDRO obtido: {resultado_api['idro']}")
                
                # Buscar itens do romaneio via GET
                print("\n" + "="*80)
                print("📥 BUSCANDO ITENS DO ROMANEIO (GET)")
                print("="*80)
                
                try:
                    dados_romaneio = api_client.get_romaneio(pedido_compra)
                    
                    if dados_romaneio and len(dados_romaneio) > 0:
                        romaneio_data = dados_romaneio[0]
                        
                        # Atualizar IDRO se ainda não tiver
                        if not romaneio.idro and 'IDRO' in romaneio_data:
                            romaneio.idro = romaneio_data['IDRO']
                            print(f"✓ IDRO obtido do GET: {romaneio_data['IDRO']}")
                        
                        # Processar itens
                        itens_api = romaneio_data.get('ITEM', [])
                        print(f"📦 {len(itens_api)} itens encontrados")
                        
                        # Salvar romaneio primeiro para ter o ID
                        db.session.add(romaneio)
                        db.session.flush()
                        
                        # Salvar itens
                        for item_data in itens_api:
                            item = RomaneioItem(
                                romaneio_id=romaneio.id,
                                codigo=item_data.get('CODIGO'),
                                descricao=item_data.get('DESCRICAO'),
                                quantidade_nf=item_data.get('QUANTIDADE_NF'),
                                quantidade_contada=item_data.get('QUANTIDADE_CONTADA')
                            )
                            db.session.add(item)
                            print(f"  ✓ Item: {item_data.get('CODIGO')} - {item_data.get('DESCRICAO')}")
                        
                        print("="*80 + "\n")
                    else:
                        print("⚠️ Nenhum item retornado pela API")
                        print("="*80 + "\n")
                        # Salvar romaneio mesmo sem itens
                        db.session.add(romaneio)
                        db.session.flush()
                        
                except Exception as e_get:
                    print(f"\n⚠️ ERRO ao buscar itens: {str(e_get)}")
                    print("="*80 + "\n")
                    # Salvar romaneio mesmo com erro ao buscar itens
                    db.session.add(romaneio)
                    db.session.flush()
                        
            except Exception as e:
                print("\n❌ ERRO NA API:")
                print("-"*80)
                print(f"🚨 Erro: {str(e)}")
                print("="*80 + "\n")
                flash(f'Erro ao chamar API: {str(e)}', 'error')
                return redirect(url_for('romaneios'))
        else:
            romaneio.idro = 999999
            print("\n" + "="*80)
            print("🧪 MODO TESTE - API NÃO CHAMADA")
            print("="*80)
            print(f"📋 Pedido: {pedido_compra}")
            print(f"📄 Nota Fiscal: {nota_fiscal}")
            print(f"🔑 Chave: {chave_acesso}")
            print(f"✓ IDRO fictício: 999999")
            print("="*80 + "\n")
            
            # No modo teste, criar itens fictícios
            db.session.add(romaneio)
            db.session.flush()
            
            # Itens de teste
            item1 = RomaneioItem(
                romaneio_id=romaneio.id,
                codigo="01.000001",
                descricao="PRODUTO TESTE A",
                quantidade_nf=100,
                quantidade_contada=100
            )
            item2 = RomaneioItem(
                romaneio_id=romaneio.id,
                codigo="01.000002",
                descricao="PRODUTO TESTE B",
                quantidade_nf=50,
                quantidade_contada=50
            )
            db.session.add(item1)
            db.session.add(item2)
        
        # Criar log
        log = RomaneioLog(
            romaneio_id=romaneio.id,
            acao='criado',
            status_novo='P',
            detalhes=f'Romaneio criado {"[MODO TESTE]" if config.MODO_TESTE else ""}',
            user_id=current_user.id
        )
        db.session.add(log)
        db.session.commit()
        
        flash('Romaneio criado com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar romaneio: {str(e)}', 'error')
    
    return redirect(url_for('romaneios'))

@app.route('/api/romaneios/<int:romaneio_id>', methods=['GET'])
@login_required
def api_get_romaneio(romaneio_id):
    """API: Buscar dados de um romaneio"""
    romaneio = Romaneio.query.get(romaneio_id)
    
    if not romaneio:
        return jsonify({'success': False, 'error': 'Romaneio não encontrado'}), 404
    
    # Incluir itens
    dados = romaneio.to_dict()
    dados['itens'] = [item.to_dict() for item in romaneio.itens]
    
    return jsonify({'success': True, 'romaneio': dados})

@app.route('/api/romaneios/<int:romaneio_id>', methods=['DELETE'])
@login_required
def api_excluir_romaneio(romaneio_id):
    """API: Excluir um romaneio"""
    try:
        romaneio = Romaneio.query.get(romaneio_id)
        
        if not romaneio:
            return jsonify({'success': False, 'error': 'Romaneio não encontrado'}), 404
        
        if not romaneio.pode_excluir():
            return jsonify({'success': False, 'error': 'Apenas romaneios pendentes sem tentativas podem ser excluídos'}), 400
        
        if not current_user.is_admin() and romaneio.created_by != current_user.id:
            return jsonify({'success': False, 'error': 'Sem permissão'}), 403
        
        db.session.delete(romaneio)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Romaneio excluído com sucesso'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/romaneios/<int:romaneio_id>/verificar', methods=['POST'])
@login_required
def api_verificar_romaneio(romaneio_id):
    """API: Forçar verificação de um romaneio específico"""
    from services.verificador_service import VerificadorService
    
    romaneio = Romaneio.query.get(romaneio_id)
    
    if not romaneio:
        return jsonify({'success': False, 'error': 'Romaneio não encontrado'}), 404
    
    if not romaneio.pode_verificar():
        return jsonify({
            'success': False,
            'error': 'Romaneio não pode ser verificado (finalizado ou máximo de tentativas atingido)'
        }), 400
    
    try:
        verificador = VerificadorService()
        resultado = verificador.verificar_romaneio(romaneio)
        
        return jsonify({
            'success': True,
            'message': 'Verificação realizada com sucesso',
            'resultado': resultado
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/romaneios/<int:romaneio_id>/status', methods=['PUT'])
@login_required
def api_atualizar_status_romaneio(romaneio_id):
    """API: Atualizar status manualmente (apenas admin)"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'error': 'Acesso negado'}), 403
    
    try:
        dados = request.get_json()
        novo_status = dados.get('status')
        observacoes = dados.get('observacoes')
        
        if not novo_status or novo_status not in config.STATUS_CHOICES:
            return jsonify({'success': False, 'error': 'Status inválido'}), 400
        
        romaneio = Romaneio.query.get(romaneio_id)
        if not romaneio:
            return jsonify({'success': False, 'error': 'Romaneio não encontrado'}), 404
        
        status_anterior = romaneio.status
        romaneio.status = novo_status
        romaneio.updated_at = datetime.utcnow()
        
        # Criar log
        log = RomaneioLog(
            romaneio_id=romaneio.id,
            acao='atualizado_manual',
            status_anterior=status_anterior,
            status_novo=novo_status,
            detalhes=observacoes or 'Status atualizado manualmente',
            user_id=current_user.id
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Status atualizado para {config.STATUS_CHOICES[novo_status]}'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/romaneios/<int:romaneio_id>/logs', methods=['GET'])
@login_required
def api_logs_romaneio(romaneio_id):
    """API: Buscar logs de um romaneio"""
    logs = RomaneioLog.query.filter_by(romaneio_id=romaneio_id)\
        .order_by(RomaneioLog.timestamp.desc()).all()
    
    return jsonify({
        'success': True,
        'logs': [log.to_dict() for log in logs]
    })

def create_admin_user():
    """Cria o usuário administrador padrão se não existir"""
    try:
        admin = User.query.filter_by(username='profectum').first()
        if not admin:
            admin = User(
                username='profectum',
                email='admin@profectum.com',
                full_name='Administrador Geral',
                role='admin'
            )
            admin.set_password('123456')
            db.session.add(admin)
            db.session.commit()
            print("👤 Usuário administrador criado: profectum / 123456")
        else:
            print("👤 Usuário administrador já existe")
    except Exception as e:
        print(f"⚠️ Erro ao verificar/criar usuário admin: {e}")
        print("💡 Execute: python init_database.py")

@app.context_processor
def inject_settings():
    """Injetar configurações do sistema nos templates"""
    return dict(
        system_settings={
            'primary_color': SystemSettings.get_setting('primary_color', '#2563eb'),
            'logo_url': SystemSettings.get_setting('logo_url', ''),
            'logo_type': SystemSettings.get_setting('logo_type', 'icon'),
            'system_name': SystemSettings.get_setting('system_name', 'RPA Profectum'),
        },
        login_settings={
            'bg_type': SystemSettings.get_setting('login_bg_type', 'color'),
            'bg_color': SystemSettings.get_setting('login_bg_color', '#f8fafc'),
            'gradient_color1': SystemSettings.get_setting('login_gradient_color1', '#2563eb'),
            'gradient_color2': SystemSettings.get_setting('login_gradient_color2', '#1d4ed8'),
            'bg_image_url': SystemSettings.get_setting('login_bg_image_url', ''),
            'height': SystemSettings.get_setting('login_height', 'normal'),
            'title': SystemSettings.get_setting('login_title', 'Bem-vindo de volta!'),
        }
    )

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        create_admin_user()
    
    print("\n" + "="*60)
    print("🚀 RPA Profectum - Sistema de Romaneios")
    print("="*60)
    print(f"Modo: {'TESTE (não chama APIs)' if config.MODO_TESTE else 'PRODUÇÃO'}")
    print(f"Painel Web: http://localhost:5000")
    print(f"\n💡 Verificador automático:")
    print(f"   Use o Agendador de Tarefas do Windows")
    print(f"   Comando: python verificador_romaneios.py --once")
    print(f"   Intervalo sugerido: {config.INTERVALO_VERIFICACAO_MINUTOS} minutos")
    print("="*60 + "\n")
    
    app.run(debug=config.FLASK_DEBUG, host='0.0.0.0', port=5000)